"""XLSX exporter — produces formatted Excel workbooks via openpyxl/pandas."""

from __future__ import annotations

import logging
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.exporters.base import BaseExporter
from app.exporters.csv_exporter import (
    BLOG_COLUMNS,
    COLLECTION_COLUMNS,
    PAGE_COLUMNS,
    REDIRECT_COLUMNS,
    SHOPIFY_PRODUCT_COLUMNS,
    URL_COLUMNS,
    _grams_from_weight,
)

logger = logging.getLogger(__name__)

# Header style constants
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_worksheet(ws: Any) -> None:
    """Apply header formatting and auto-fit column widths."""
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), start=1):
        max_length = 0
        for cell in col_cells:
            try:
                length = len(str(cell.value or ""))
            except Exception:
                length = 0
            max_length = max(max_length, length)
        # Clamp between 10 and 60 characters
        adjusted = min(max(max_length + 3, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    ws.freeze_panes = "A2"


def _write_styled_xlsx(df: pd.DataFrame, path: str, sheet_name: str = "Sheet1") -> None:
    """Write a DataFrame to an xlsx file with styled headers."""
    df.to_excel(path, index=False, sheet_name=sheet_name, engine="openpyxl")
    wb = load_workbook(path)
    _style_worksheet(wb[sheet_name])
    wb.save(path)
    wb.close()


class XLSXExporter(BaseExporter):
    """Export data as styled XLSX workbooks."""

    # ------------------------------------------------------------------
    # Products
    # ------------------------------------------------------------------

    async def export_products(
        self, products: List[Dict[str, Any]], variants: List[Dict[str, Any]]
    ) -> str:
        variants_by_product: Dict[int, List[Dict[str, Any]]] = {}
        for v in variants:
            pid = v.get("product_id")
            if pid is not None:
                variants_by_product.setdefault(pid, []).append(v)

        rows: list[dict[str, Any]] = []
        for product in products:
            pid = product.get("id")
            handle = product.get("handle", "")
            image_urls: list[str] = product.get("image_urls") or []
            prod_variants = variants_by_product.get(pid, [])

            base: dict[str, Any] = {
                "Handle": handle,
                "Title": product.get("title", ""),
                "Body (HTML)": product.get("description_html", ""),
                "Vendor": product.get("vendor", ""),
                "Product Category": "",
                "Type": product.get("product_type", ""),
                "Tags": product.get("tags", ""),
                "Published": "TRUE",
                "SEO Title": product.get("seo_title", ""),
                "SEO Description": product.get("seo_description", ""),
                "Status": product.get("status", "active"),
            }

            if prod_variants:
                for idx, var in enumerate(prod_variants):
                    row = dict(base) if idx == 0 else {"Handle": handle}
                    row.update(self._variant_fields(var))
                    if idx == 0 and image_urls:
                        row["Image Src"] = image_urls[0]
                        row["Image Position"] = 1
                        row["Image Alt Text"] = product.get("title", "")
                    rows.append(row)

                for img_idx, img_url in enumerate(image_urls[1:], start=2):
                    rows.append(
                        {
                            "Handle": handle,
                            "Image Src": img_url,
                            "Image Position": img_idx,
                            "Image Alt Text": product.get("title", ""),
                        }
                    )
            else:
                base.update(
                    {
                        "Option1 Name": "Title",
                        "Option1 Value": "Default Title",
                        "Variant SKU": product.get("sku", ""),
                        "Variant Grams": 0,
                        "Variant Inventory Tracker": "shopify",
                        "Variant Inventory Qty": 0,
                        "Variant Inventory Policy": "deny",
                        "Variant Fulfillment Service": "manual",
                        "Variant Price": product.get("price", "0.00"),
                        "Variant Compare At Price": product.get(
                            "compare_at_price", ""
                        ),
                        "Variant Requires Shipping": "TRUE",
                        "Variant Taxable": "TRUE",
                        "Variant Barcode": product.get("barcode", ""),
                    }
                )
                if image_urls:
                    base["Image Src"] = image_urls[0]
                    base["Image Position"] = 1
                    base["Image Alt Text"] = product.get("title", "")
                rows.append(base)

                for img_idx, img_url in enumerate(image_urls[1:], start=2):
                    rows.append(
                        {
                            "Handle": handle,
                            "Image Src": img_url,
                            "Image Position": img_idx,
                            "Image Alt Text": product.get("title", ""),
                        }
                    )

        df = pd.DataFrame(rows, columns=SHOPIFY_PRODUCT_COLUMNS).fillna("")
        path = self._filepath("products", "xlsx")
        _write_styled_xlsx(df, path, sheet_name="Products")

        # Add data validations
        wb = load_workbook(path)
        ws = wb["Products"]
        status_col = SHOPIFY_PRODUCT_COLUMNS.index("Status") + 1
        status_letter = get_column_letter(status_col)
        dv = DataValidation(
            type="list",
            formula1='"active,draft,archived"',
            allow_blank=True,
        )
        dv.error = "Status must be active, draft, or archived"
        dv.errorTitle = "Invalid Status"
        dv.add(f"{status_letter}2:{status_letter}{len(rows) + 1}")
        ws.add_data_validation(dv)

        published_col = SHOPIFY_PRODUCT_COLUMNS.index("Published") + 1
        pub_letter = get_column_letter(published_col)
        dv_pub = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
        )
        dv_pub.add(f"{pub_letter}2:{pub_letter}{len(rows) + 1}")
        ws.add_data_validation(dv_pub)

        wb.save(path)
        wb.close()

        logger.info("Exported %d product rows to %s", len(df), path)
        return path

    @staticmethod
    def _variant_fields(var: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "Option1 Name": var.get("option1_name", ""),
            "Option1 Value": var.get("option1_value", ""),
            "Option2 Name": var.get("option2_name", ""),
            "Option2 Value": var.get("option2_value", ""),
            "Option3 Name": var.get("option3_name", ""),
            "Option3 Value": var.get("option3_value", ""),
            "Variant SKU": var.get("sku", ""),
            "Variant Grams": _grams_from_weight(
                var.get("weight"), var.get("weight_unit")
            ),
            "Variant Inventory Tracker": "shopify",
            "Variant Inventory Qty": var.get("inventory_qty", 0),
            "Variant Inventory Policy": "deny",
            "Variant Fulfillment Service": "manual",
            "Variant Price": var.get("price", "0.00"),
            "Variant Compare At Price": var.get("compare_at_price", ""),
            "Variant Requires Shipping": "TRUE",
            "Variant Taxable": "TRUE",
            "Variant Barcode": var.get("barcode", ""),
        }

    # ------------------------------------------------------------------
    # Collections
    # ------------------------------------------------------------------

    async def export_collections(self, collections: List[Dict[str, Any]]) -> str:
        rows: list[dict[str, Any]] = []
        for col in collections:
            handles = col.get("product_handles") or []
            rows.append(
                {
                    "Title": col.get("title", ""),
                    "Handle": col.get("handle", ""),
                    "Body HTML": col.get("description_html", ""),
                    "Sort Order": col.get("sort_order", ""),
                    "Product Handles": ";".join(handles)
                    if isinstance(handles, list)
                    else str(handles),
                    "SEO Title": col.get("seo_title", ""),
                    "SEO Description": col.get("seo_description", ""),
                }
            )
        df = pd.DataFrame(rows, columns=COLLECTION_COLUMNS).fillna("")
        path = self._filepath("collections", "xlsx")
        _write_styled_xlsx(df, path, sheet_name="Collections")
        logger.info("Exported %d collections to %s", len(df), path)
        return path

    # ------------------------------------------------------------------
    # Pages
    # ------------------------------------------------------------------

    async def export_pages(self, pages: List[Dict[str, Any]]) -> str:
        rows: list[dict[str, Any]] = []
        for page in pages:
            rows.append(
                {
                    "Title": page.get("title", ""),
                    "Handle": page.get("handle", ""),
                    "Body HTML": page.get("body_html", ""),
                    "Published": "TRUE" if page.get("published", True) else "FALSE",
                    "SEO Title": page.get("seo_title", ""),
                    "SEO Description": page.get("seo_description", ""),
                }
            )
        df = pd.DataFrame(rows, columns=PAGE_COLUMNS).fillna("")
        path = self._filepath("pages", "xlsx")
        _write_styled_xlsx(df, path, sheet_name="Pages")
        logger.info("Exported %d pages to %s", len(df), path)
        return path

    # ------------------------------------------------------------------
    # Blogs
    # ------------------------------------------------------------------

    async def export_blogs(self, blogs: List[Dict[str, Any]]) -> str:
        rows: list[dict[str, Any]] = []
        for blog in blogs:
            rows.append(
                {
                    "Blog Title": blog.get("blog_title", ""),
                    "Title": blog.get("title", ""),
                    "Handle": blog.get("handle", ""),
                    "Author": blog.get("author", ""),
                    "Body HTML": blog.get("body_html", ""),
                    "Tags": blog.get("tags", ""),
                    "Featured Image": blog.get("featured_image", ""),
                    "Published At": blog.get("published_at", ""),
                    "SEO Title": blog.get("seo_title", ""),
                    "SEO Description": blog.get("seo_description", ""),
                }
            )
        df = pd.DataFrame(rows, columns=BLOG_COLUMNS).fillna("")
        path = self._filepath("blogs", "xlsx")
        _write_styled_xlsx(df, path, sheet_name="Blog Posts")
        logger.info("Exported %d blog posts to %s", len(df), path)
        return path

    # ------------------------------------------------------------------
    # URLs
    # ------------------------------------------------------------------

    async def export_urls(self, urls: List[Dict[str, Any]]) -> str:
        rows: list[dict[str, Any]] = []
        for u in urls:
            rows.append(
                {
                    "URL": u.get("url", ""),
                    "Status Code": u.get("status_code", ""),
                    "Content Type": u.get("content_type", ""),
                    "Canonical URL": u.get("canonical_url", ""),
                    "Meta Title": u.get("meta_title", ""),
                    "Meta Description": u.get("meta_description", ""),
                    "Page Type": u.get("page_type", ""),
                }
            )
        df = pd.DataFrame(rows, columns=URL_COLUMNS).fillna("")
        path = self._filepath("urls", "xlsx")
        _write_styled_xlsx(df, path, sheet_name="URLs")
        logger.info("Exported %d URL records to %s", len(df), path)
        return path

    # ------------------------------------------------------------------
    # Redirects
    # ------------------------------------------------------------------

    async def export_redirects(self, redirects: List[Dict[str, Any]]) -> str:
        rows: list[dict[str, Any]] = []
        for r in redirects:
            rows.append(
                {
                    "Old URL": r.get("old_url", "") or r.get("url", ""),
                    "New URL": r.get("new_url", "") or r.get("redirect_to", ""),
                }
            )
        df = pd.DataFrame(rows, columns=REDIRECT_COLUMNS).fillna("")
        path = self._filepath("redirects", "xlsx")
        _write_styled_xlsx(df, path, sheet_name="Redirects")
        logger.info("Exported %d redirects to %s", len(df), path)
        return path
